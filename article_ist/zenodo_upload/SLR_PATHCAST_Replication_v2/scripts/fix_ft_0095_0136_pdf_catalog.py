#!/usr/bin/env python3
"""Fix duplicate PDF catalog entry for ft_0095 / ft_0136.

The two review_ids pointed to the same file (MD5 66ef2b93…), which is the
ICSESS 2015 paper by Jiang et al. (AADL/FMEA/Markov) — correct for ft_0095,
wrong for ft_0136 (Nigro et al. 2019 Springer chapter, not OA).

Actions:
  1. Move ft_0136 duplicate from ok/ to bad_match/
  2. Clear local_pdf_path for ft_0136 in blind review CSV + XLSX
  3. Append issue record to ft_pdf_catalog_issues.csv
  4. Verify ft_0095 PDF remains in ok/ and path unchanged

Usage:
    python scripts/fix_ft_0095_0136_pdf_catalog.py
    python scripts/fix_ft_0095_0136_pdf_catalog.py --dry-run
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
HUMAN_DIR = REPO / "results/human_validation"
OK_DIR = HUMAN_DIR / "ft_pdfs_local/ok"
BAD_DIR = HUMAN_DIR / "ft_pdfs_local/bad_match"
ISSUES_CSV = HUMAN_DIR / "ft_pdf_catalog_issues.csv"
SHEET_CSV = HUMAN_DIR / "ft_qa_extraction_blind_review_sheet.csv"
SHEET_XLSX = HUMAN_DIR / "ft_qa_extraction_blind_review_sheet.xlsx"

FT_0095 = "ft_0095"
FT_0136 = "ft_0136"
SHARED_MD5 = "66ef2b9369ec2422b778d044ad74ebe9"
FT_0095_PDF = OK_DIR / "ft_0095_model_driven_safety_modeling_and_analysis_of_embedde.pdf"
FT_0136_PDF = OK_DIR / "ft_0136_formal_modeling_and_analysis_of_probabilistic_real_t.pdf"
FT_0136_BAD = BAD_DIR / "ft_0136_WRONG_DUPLICATE_OF_ft_0095.pdf"
FT_0095_REL = "results/human_validation/ft_pdfs_local/ok/ft_0095_model_driven_safety_modeling_and_analysis_of_embedde.pdf"

ISSUE_COLS = [
    "detected_at",
    "review_id",
    "issue_type",
    "details",
    "action_taken",
    "local_pdf_path_before",
    "local_pdf_path_after",
]


def _md5(path: Path) -> str:
    return hashlib.md5(path.read_bytes()).hexdigest()


def _rel(path: Path) -> str:
    try:
        return str(path.relative_to(REPO))
    except ValueError:
        return str(path)


def update_csv(dry_run: bool) -> None:
    rows: list[dict] = []
    with SHEET_CSV.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        for row in reader:
            rid = row.get("review_id")
            if rid == FT_0136:
                row["local_pdf_path"] = ""
            elif rid == FT_0095 and not (row.get("local_pdf_path") or "").strip():
                row["local_pdf_path"] = FT_0095_REL
            rows.append(row)
    if dry_run:
        hit = [r for r in rows if r.get("review_id") == FT_0136]
        print("[dry-run] CSV ft_0136 local_pdf_path ->", repr(hit[0]["local_pdf_path"]) if hit else "missing")
        return
    with SHEET_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def update_xlsx(dry_run: bool) -> None:
    wb = openpyxl.load_workbook(SHEET_XLSX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    if "local_pdf_path" not in idx:
        raise KeyError("local_pdf_path column missing in xlsx")
    for row in ws.iter_rows(min_row=2):
        if row[idx["review_id"]].value == FT_0136:
            if dry_run:
                print("[dry-run] XLSX ft_0136 local_pdf_path -> ''")
            else:
                row[idx["local_pdf_path"]].value = ""
    if not dry_run:
        wb.save(SHEET_XLSX)


def append_issue(dry_run: bool, before: str, after: str) -> None:
    if ISSUES_CSV.exists():
        with ISSUES_CSV.open(newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("review_id") == FT_0136 and row.get("issue_type") == "duplicate_pdf_wrong_match":
                    print("Issue record already present in ft_pdf_catalog_issues.csv")
                    return
    record = {
        "detected_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "review_id": FT_0136,
        "issue_type": "duplicate_pdf_wrong_match",
        "details": (
            f"ft_0136 shared MD5 {SHARED_MD5} with ft_0095 (Jiang et al. ICSESS 2015). "
            "Expected Nigro et al. 2019 (DOI 10.1007/978-981-15-0637-6_5), not OA."
        ),
        "action_taken": "moved duplicate to bad_match/; cleared local_pdf_path for ft_0136",
        "local_pdf_path_before": before,
        "local_pdf_path_after": after,
    }
    if dry_run:
        print("[dry-run] issue record:", record)
        return
    write_header = not ISSUES_CSV.exists()
    with ISSUES_CSV.open("a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ISSUE_COLS)
        if write_header:
            writer.writeheader()
        writer.writerow(record)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not FT_0095_PDF.exists():
        raise FileNotFoundError(f"Missing expected ft_0095 PDF: {FT_0095_PDF}")
    if _md5(FT_0095_PDF) != SHARED_MD5:
        raise ValueError(f"ft_0095 MD5 mismatch: {_md5(FT_0095_PDF)}")

    before_path = _rel(FT_0136_PDF) if FT_0136_PDF.exists() else ""

    if FT_0136_PDF.exists():
        if _md5(FT_0136_PDF) != SHARED_MD5:
            raise ValueError("ft_0136 PDF exists but MD5 differs from expected duplicate")
        if args.dry_run:
            print(f"[dry-run] would move {FT_0136_PDF} -> {FT_0136_BAD}")
        else:
            BAD_DIR.mkdir(parents=True, exist_ok=True)
            shutil.move(str(FT_0136_PDF), str(FT_0136_BAD))
            print(f"Moved duplicate to {_rel(FT_0136_BAD)}")
    else:
        print("ft_0136 PDF already absent from ok/")

    update_csv(args.dry_run)
    update_xlsx(args.dry_run)
    append_issue(args.dry_run, before_path, "")

    print(f"ft_0095 kept: {_rel(FT_0095_PDF)} (MD5 verified)")
    print("ft_0136 local_pdf_path cleared — manual download required (Springer, not OA)")


if __name__ == "__main__":
    main()
